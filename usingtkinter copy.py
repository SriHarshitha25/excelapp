import tkinter as tk
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
import openpyxl
import datetime
import os

class StoreApp:
    def __init__(self, root):
        self.root = root
        self.root.geometry("800x800")
        self.root.title('Store Supplies')

        # StringVar to store the selected date
        self.date_var = tk.StringVar()

        # Date widget
        ttk.Label(root, text="Requirement Date:").grid(row=3, column=0, padx=10, pady=10)
        date_entry = Entry(root, textvariable=self.date_var)
        date_entry.grid(row=3, column=1, padx=15)

        # Load data from Excel file
        self.file_path = "store_supplies.xlsx"
        if not os.path.exists(self.file_path):
            self.create_initial_excel()

        self.data = self.load_excel_data(self.file_path)

        # Defining the variables
        self.material_var = tk.StringVar()
        self.description_var = tk.StringVar()

        # Label for Material Name
        ttk.Label(root, text="Material Name:").grid(row=0, column=0, padx=10, pady=10)

        # Entry widget for typing Material Name
        self.material_entry = Entry(root, textvariable=self.material_var)
        self.material_entry.grid(row=0, column=1, padx=10, pady=10)
        self.material_entry.bind("<Return>", self.show_material_details)

        # Quantity Entry
        ttk.Label(root, text="Quantity:").grid(row=1, column=0, padx=10, pady=10)
        self.quantity_var = tk.IntVar()
        self.quantity_entry = Entry(root, textvariable=self.quantity_var)
        self.quantity_entry.grid(row=1, column=1, padx=10, pady=10)

        # Add and Subtract Buttons
        add_button = ttk.Button(root, text="Add", command=self.add_quantity)
        add_button.grid(row=2, column=0, padx=10, pady=10)

        subtract_button = ttk.Button(root, text="Subtract", command=self.subtract_quantity)
        subtract_button.grid(row=2, column=1, padx=10, pady=10)

    def create_initial_excel(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Store Supplies"
        sheet['A1'] = "Material"
        sheet['B1'] = "Code"
        sheet['C1'] = "Qty"
        workbook.save(self.file_path)

    def load_excel_data(self, file_path):
        data = {}
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            for row in sheet.iter_rows(values_only=True):
                if len(row) == 3:
                    material, code, qty = row
                    data[material] = {'code': code, 'qty': qty}

            workbook.close()
        except Exception as e:
            print(f"Error loading data from Excel: {e}")

        return data

    def show_material_details(self, event):
        material_name = self.material_var.get()
        if material_name in self.data:
            material_data = self.data[material_name]
            material_code = material_data['code']
            material_qty = material_data['qty']
            material_info = f"Material Name: {material_name}\nMaterial Code: {material_code}\nQuantity: {material_qty}"
            messagebox.showinfo("Material Details", material_info)
        else:
            messagebox.showinfo("Material Details", "Material not found")

    def update_excel(self, sheet_name, log_operation, material_name, date, quantity_change):
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            if sheet_name not in workbook.sheetnames:
                workbook.create_sheet(sheet_name)
                sheet = workbook[sheet_name]
                sheet['A1'] = "Material Name"
                sheet['B1'] = "Date"
                sheet['C1'] = "Quantity Change"
            else:
                sheet = workbook[sheet_name]

            sheet.append([material_name, date, quantity_change])

            # Update the main sheet with quantity change
            main_sheet = workbook.active
            for row in range(2, main_sheet.max_row + 1):
                material = main_sheet.cell(row=row, column=1).value
                if material == material_name:
                    qty = main_sheet.cell(row=row, column=3).value + quantity_change
                    main_sheet.cell(row=row, column=3, value=qty)
                    break

            workbook.save(self.file_path)
            workbook.close()
        except Exception as e:
            print(f"Error updating Excel data: {e}")

    def add_quantity(self):
        material_name = self.material_var.get()
        date = self.date_var.get()
        quantity_change = self.quantity_var.get()
        if not material_name or not date or quantity_change <= 0:
            messagebox.showinfo("Error", "Please enter valid data.")
            return

        self.update_excel("Add_Log", "Add", material_name, date, quantity_change)
        self.material_var.set("")
        self.date_var.set("")
        self.quantity_var.set(0)

    def subtract_quantity(self):
        material_name = self.material_var.get()
        date = self.date_var.get()
        quantity_change = -self.quantity_var.get()
        if not material_name or not date or quantity_change >= 0:
            messagebox.showinfo("Error", "Please enter valid data.")
            return

        self.update_excel("Subtract_Log", "Subtract", material_name, date, quantity_change)
        self.material_var.set("")
        self.date_var.set("")
        self.quantity_var.set(0)

# Create the Tkinter window and run the app
root = tk.Tk()
app = StoreApp(root)
root.mainloop()

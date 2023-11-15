import tkinter as tk
from tkinter import Entry, ttk
import openpyxl
from openpyxl import load_workbook, Workbook
from tkinter import *
import tkinter.ttk as ttk

#i figured out the search option babeee


# Load the xlsx file, then store the value of each column in the "elements" list
wb = load_workbook(filename=r"C:\Users\Chinnu\Downloads\exampleapp\testdata.xlsx")
ws = wb['Sheet1']
wBook = load_workbook('storingfile.xlsx')
sheet = wBook.active

m_row = ws.max_row
m_col=  ws.max_column
print("total rows of the excel: ",m_row)
MaterialDescription = ws['A2':'A10']
elements = []  
codeval = ''     

#to get the list of column values
for cell in MaterialDescription:
    for x in cell:
        y = x.value
        elements.append(y)
        print(y)


#search function
def check_input(event):
    value2 = event.widget.get()

    if value2 == '':
        combodata['values'] = elements
    else:
        data = []
        for item in elements:
            if value2.lower() in item.lower():
                data.append(item)

        combodata['values'] = data

        print("selected value=", combodata.get(), "index=",combodata.current())

        main_sheet = wb.active
        for row in range(2, main_sheet.max_row + 1):
            material = main_sheet.cell(row=row, column=1).value
            if material == combodata.get():
                codeval = main_sheet.cell(row=row, column=2).value
                print(main_sheet.cell(row=row, column=2).value)
                code.insert("end-1c", codeval)
                break         
        

        data1 = [combodata.get()]
        sheet.append(data1)
        wBook.save('storingfile.xlsx')


# Tkinter stuff
win = Tk()
clicked = StringVar()

#label and combobox, binding
ttk.Label(text="Material Description:").grid(row=1, column=0, padx=10, pady=10)
combodata = ttk.Combobox(win, values=elements)
combodata.grid(row=1, column=1, padx=10, pady=10)
combodata['values'] = elements
combodata.bind('<KeyRelease>', check_input)
wBook.save('storingfile.xlsx')

ttk.Label(text="Material Code:").grid(row=2, column=0, padx=10, pady=10)
code = Text(height=1, width=18)
code.grid(row=2, column=1, padx=10, pady=10)

win.mainloop()
